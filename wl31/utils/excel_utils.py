# wl31/utils/excel_utils.py
import openpyxl
from openpyxl.utils import get_column_letter

def export_to_excel(data, headers, filepath):
    """
    将数据导出到 Excel 文件。

    Args:
        data (list of lists): 要导出的数据，每一行是一个列表。
        headers (list of str): Excel 文件的表头。
        filepath (str): 保存文件的路径。
    """
    if not filepath:
        return False
    
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # 写入表头
        sheet.append(headers)
        
        # 写入数据
        for row_data in data:
            sheet.append(row_data)
            
        # 调整列宽
        for i, column_cells in enumerate(sheet.columns):
            max_length = 0
            column = get_column_letter(i + 1)
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        workbook.save(filepath)
        return True
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        return False

def import_from_excel(filepath, column_mapping):
    """
    从 Excel 文件导入数据。

    Args:
        filepath (str): Excel 文件的路径。
        column_mapping (dict): 列索引到字段名的映射。
                                e.g., {0: 'name', 1: 'age'}

    Returns:
        list of dict: 包含导入数据的字典列表，或在出错时返回 None。
    """
    if not filepath:
        return None
    
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        
        data = []
        # 从第二行开始迭代，跳过标题行
        for row_index, row in enumerate(sheet.iter_rows(min_row=2)):
            row_data = {}
            for col_index, field_name in column_mapping.items():
                cell_value = row[col_index].value
                row_data[field_name] = cell_value
            data.append(row_data)
            
        return data
    except Exception as e:
        print(f"Error importing from Excel: {e}")
        return None

def import_students_from_excel(filepath):
    """
    专门用于从 Excel 导入学生信息的函数。
    列顺序: 学院, 班级, 姓名, 性别, 入学年份, 身份证号, 联系方式
    """
    column_mapping = {
        0: 'department',
        1: 'class',
        2: 'name',
        3: 'gender',
        4: 'enrollment_year',
        5: 'id_card',
        6: 'contact'
    }
    return import_from_excel(filepath, column_mapping)

def import_teachers_from_excel(filepath):
    """
    专门用于从 Excel 导入教师信息的函数。
    列顺序: 学院, 姓名, 性别, 职称, 身份证号, 联系方式
    """
    column_mapping = {
        0: 'department',
        1: 'name',
        2: 'gender',
        3: 'title',
        4: 'id_card',
        5: 'contact'
    }
    return import_from_excel(filepath, column_mapping)